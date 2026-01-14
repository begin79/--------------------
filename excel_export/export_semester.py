"""
Утилита для экспорта расписания за семестр в Excel.

Запуск:
    python -m excel_export.export_semester --query "ИС1-227-ОТ" --mode student
"""
from __future__ import annotations

import argparse
import asyncio
import datetime as dt
import io
import logging
import os
import re
import sys
import time
import zipfile
from collections import OrderedDict
from importlib import import_module
from pathlib import Path
from typing import Dict, Iterable, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    from telegram import Bot
    from telegram.error import TelegramError
except Exception:  # pragma: no cover
    Bot = None  # type: ignore
    TelegramError = Exception  # type: ignore

ROOT_DIR = Path(__file__).resolve().parents[1]
if str(ROOT_DIR) not in sys.path:
    sys.path.insert(0, str(ROOT_DIR))

os.environ.setdefault("BOT_TOKEN", os.environ.get("BOT_TOKEN", "excel-export-placeholder"))

from app.constants import API_TYPE_GROUP, API_TYPE_TEACHER  # noqa: E402
from app.schedule import get_schedule_structured  # noqa: E402

logger = logging.getLogger("excel_export")
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")

WEEKDAY_NAMES = [
    "Понедельник",
    "Вторник",
    "Среда",
    "Четверг",
    "Пятница",
    "Суббота",
    "Воскресенье",
]

SEMESTER_PRESETS = {
    "autumn": {
        "label": "Осенний семестр",
        "start": (9, 1),
        "end": (12, 31),
    },
    "spring": {
        "label": "Весенний семестр",
        "start": (1, 1),
        "end": (4, 30),
    },
}

try:
    config = import_module("excel_export.config")
except ModuleNotFoundError:
    config = None

CONFIG_BOT_TOKEN = getattr(config, "TELEGRAM_BOT_TOKEN", os.getenv("EXPORT_BOT_TOKEN"))
CONFIG_CHAT_ID = getattr(config, "TELEGRAM_CHAT_ID", os.getenv("EXPORT_CHAT_ID"))


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Экспорт расписания за семестр в Excel")
    parser.add_argument("--query", required=True, help="Название группы или преподавателя")
    parser.add_argument(
        "--mode",
        choices=("student", "teacher"),
        default="student",
        help="Режим экспорта: группа (student) или преподаватель (teacher)",
    )
    parser.add_argument(
        "--semester",
        choices=("autumn", "spring"),
        help="Семестр: autumn (сентябрь-декабрь) или spring (январь-апрель). По умолчанию определяется по текущей дате.",
    )
    parser.add_argument(
        "--year",
        type=int,
        help="Год семестра (например 2025). Если не указан, определяется автоматически.",
    )
    parser.add_argument("--start", type=str, help="Переопределить дату начала (формат YYYY-MM-DD)")
    parser.add_argument("--end", type=str, help="Переопределить дату окончания (формат YYYY-MM-DD)")
    try:
        from app.config import EXPORTS_DIR
        default_output = EXPORTS_DIR
    except ImportError:
        default_output = "exports"

    parser.add_argument(
        "--output",
        type=str,
        default=default_output,
        help=f"Каталог для сохранения Excel-файла (по умолчанию {default_output})",
    )
    parser.add_argument("--filename", type=str, help="Имя файла (по умолчанию генерируется автоматически)")
    parser.add_argument(
        "--send",
        action="store_true",
        help="Отправить результат в Telegram (нужен токен и chat_id)",
    )
    parser.add_argument("--bot-token", type=str, help="Токен Telegram-бота для отправки файла")
    parser.add_argument("--chat-id", type=str, help="ID чата для отправки файла")
    parser.add_argument("--silent", action="store_true", help="Отключить информационные логи")
    parser.add_argument(
        "--split-groups",
        action="store_true",
        help="Для преподавателя дополнительно сохранить отдельный Excel-файл на каждую группу",
    )
    return parser.parse_args()


def resolve_semester_bounds(
    semester: Optional[str], year: Optional[int], start: Optional[str], end: Optional[str]
) -> Tuple[dt.date, dt.date, str]:
    today = dt.date.today()
    if semester is None:
        if today.month >= 9:
            semester = "autumn"
        elif today.month <= 4:
            semester = "spring"
        else:
            semester = "autumn"

    preset = SEMESTER_PRESETS[semester]
    if year is None:
        if semester == "autumn":
            year = today.year if today.month >= 9 else today.year - 1
        else:  # spring
            year = today.year if today.month <= 4 else today.year + 1

    start_date = dt.date(year, preset["start"][0], preset["start"][1])
    end_date = dt.date(year, preset["end"][0], preset["end"][1])

    label = f"{preset['label']} {year}"
    if start:
        start_date = _parse_date(start)
        label = "Произвольный период"
    if end:
        end_date = _parse_date(end)
        label = "Произвольный период"
    if start_date > end_date:
        raise ValueError("Дата начала позже даты окончания")
    return start_date, end_date, label


def _parse_date(value: str) -> dt.date:
    return dt.datetime.strptime(value, "%Y-%m-%d").date()


async def fetch_semester_schedule(
    query: str,
    entity_type: str,
    start_date: dt.date,
    end_date: dt.date,
    bot=None,
) -> "OrderedDict[dt.date, Dict]":
    valid_days: List[dt.date] = []
    cursor = start_date
    while cursor <= end_date:
        if cursor.weekday() != 6:  # воскресенье пропускаем
            valid_days.append(cursor)
        cursor += dt.timedelta(days=1)

    total_days = len(valid_days)
    logger.info(
        "📅 Собираю расписание для '%s' (%s) с %s по %s (%d учебных дней)",
        query,
        "группа" if entity_type == API_TYPE_GROUP else "преподаватель",
        start_date.strftime("%d.%m.%Y"),
        end_date.strftime("%d.%m.%Y"),
        total_days,
    )

    semaphore = asyncio.Semaphore(8)
    results: Dict[dt.date, Dict] = {}
    start_ts = time.perf_counter()

    async def fetch_day(date_obj: dt.date):
        date_str = date_obj.strftime("%Y-%m-%d")
        async with semaphore:
            structured, err = await get_schedule_structured(date_str, query, entity_type, bot=bot)
        if err:
            logger.debug("Нет расписания на %s: %s", date_str, err)
            return
        if not structured or not structured.get("pairs"):
            return
        actual_date_str = structured.get("date_iso", date_str)
        try:
            actual_date = dt.datetime.strptime(actual_date_str, "%Y-%m-%d").date()
        except ValueError:
            actual_date = date_obj
        results[actual_date] = {
            "weekday": structured.get("weekday") or WEEKDAY_NAMES[actual_date.weekday()],
            "pairs": structured.get("pairs", []),
        }

    await asyncio.gather(*(fetch_day(day) for day in valid_days))
    ordered = OrderedDict(sorted(results.items(), key=lambda item: item[0]))
    duration = time.perf_counter() - start_ts
    logger.info(
        "✅ Найдено %d учебных дней с расписанием (%.1f c)",
        len(ordered),
        duration,
    )
    return ordered


def build_excel_workbook(
    entity_name: str,
    mode: str,
    semester_label: str,
    data: "OrderedDict[dt.date, Dict]",
) -> Tuple[
    Workbook,
    Dict[str, List[List[str]]],
    Dict[str, List[List[str]]],
    float,
    Dict[str, float],
    Dict[str, float],
]:
    wb = Workbook()
    ws = wb.active
    sheet_name = _sanitize_sheet_name(entity_name)
    ws.title = sheet_name

    entity_label = "Группа" if mode == "student" else "Преподаватель"
    ws.merge_cells("A1:I1")
    ws["A1"] = f"Расписание за семестр ({semester_label})"
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")

    ws.merge_cells("A2:I2")
    ws["A2"] = f"{entity_label}: {entity_name}"
    ws["A2"].font = Font(bold=True, size=13)
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Дата",
        "День недели",
        "№ пары",
        "Время",
        "Дисциплина",
        "Преподаватель",
        "Аудитория",
        "Группы / подгруппа",
        "Комментарий",
    ]
    ws.append([])
    ws.append(headers)

    header_row = ws[4]
    header_fill = PatternFill("solid", fgColor="E8F5E9")
    thin_side = Side(style="thin", color="CCCCCC")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    per_group_rows: Dict[str, List[List[str]]] = {}
    per_teacher_rows: Dict[str, List[List[str]]] = {}
    per_group_hours: Dict[str, float] = {}
    per_teacher_hours: Dict[str, float] = {}
    total_hours = 0.0

    for date_obj, info in data.items():
        weekday = info.get("weekday") or WEEKDAY_NAMES[date_obj.weekday()]
        pairs = info.get("pairs", [])
        if not pairs:
            continue

        pair_counter = 0
        last_time = ""
        for pair in pairs:
            time_slot = pair.get("time", "-")
            if time_slot != last_time:
                pair_counter += 1
                last_time = time_slot

            teacher = pair.get("teacher") or pair.get("fio") or ""
            auditorium = pair.get("auditorium") or pair.get("room") or ""

            groups_list = _extract_groups(pair)
            groups_str = ", ".join(groups_list)

            subgroup = pair.get("subgroup")
            if subgroup:
                groups_str = f"{groups_str} ({subgroup})" if groups_str else str(subgroup)

            comment_parts = [
                pair.get("type"),
                pair.get("comment"),
                pair.get("note"),
            ]
            comment = "; ".join(str(part) for part in comment_parts if part)

            duration = _calculate_pair_duration(time_slot)
            total_hours += duration

            row = [
                date_obj.strftime("%d.%m.%Y"),
                weekday,
                pair_counter,
                time_slot,
                pair.get("subject", "-"),
                teacher,
                auditorium,
                groups_str,
                comment,
            ]
            ws.append(row)

            if mode == "teacher" and groups_list:
                for grp in groups_list:
                    per_group_rows.setdefault(grp, []).append(
                        [
                            date_obj.strftime("%d.%m.%Y"),
                            weekday,
                            pair_counter,
                            time_slot,
                            pair.get("subject", "-"),
                            auditorium,
                            comment,
                        ]
                    )
                    per_group_hours[grp] = per_group_hours.get(grp, 0.0) + duration
            if mode == "student":
                teacher_name = teacher.strip()
                if teacher_name:
                    per_teacher_rows.setdefault(teacher_name, []).append(
                        [
                            date_obj.strftime("%d.%m.%Y"),
                            weekday,
                            pair_counter,
                            time_slot,
                            pair.get("subject", "-"),
                            teacher,
                            auditorium,
                            comment,
                        ]
                    )
                    per_teacher_hours[teacher_name] = per_teacher_hours.get(teacher_name, 0.0) + duration

    ws.freeze_panes = "A5"
    _auto_fit_columns(ws, max_width=45)
    if total_hours > 0:
        ws.append([])
        summary_row = [
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            "",
            f"Итого: {total_hours:.1f} часов",
        ]
        ws.append(summary_row)
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)

    if mode == "teacher" and per_group_rows:
        _add_group_sheets(wb, per_group_rows, per_group_hours)
    if mode == "student" and per_teacher_rows:
        _add_teacher_sheets(wb, per_teacher_rows, per_teacher_hours)

    return wb, per_group_rows, per_teacher_rows, total_hours, per_group_hours, per_teacher_hours


def _auto_fit_columns(ws, max_width: int = 40) -> None:
    for column_cells in ws.columns:
        letter = get_column_letter(column_cells[0].column)
        max_length = 0
        for cell in column_cells:
            value = cell.value or ""
            max_length = max(max_length, len(str(value)))
        ws.column_dimensions[letter].width = min(max_width, max_length + 2)


def _sanitize_sheet_name(value: str) -> str:
    invalid = set('[]:*?/\\')
    cleaned = "".join("_" if ch in invalid else ch for ch in value)
    if len(cleaned) > 31:
        cleaned = cleaned[:31]
    return cleaned or "Sheet1"


def _sanitize_filename(value: str) -> str:
    cleaned = re.sub(r'[\\/*?:"<>|]', "_", value).strip()
    return cleaned or "export"


def _extract_groups(pair: Dict) -> List[str]:
    groups_data = pair.get("groups") or []
    if isinstance(groups_data, str):
        result = [groups_data.strip()] if groups_data.strip() else []
    elif isinstance(groups_data, Iterable):
        result = [str(item).strip() for item in groups_data if str(item).strip()]
    else:
        result = []
    return result


def _calculate_pair_duration(time_slot: str) -> float:
    if not time_slot:
        return 1.5
    match = re.search(r"(\d{1,2})[:.](\d{2})\s*[-–]\s*(\d{1,2})[:.](\d{2})", time_slot)
    if not match:
        return 1.5
    start_h, start_m, end_h, end_m = map(int, match.groups())
    start = dt.timedelta(hours=start_h, minutes=start_m)
    end = dt.timedelta(hours=end_h, minutes=end_m)
    duration = (end - start).total_seconds() / 3600
    if duration <= 0:
        duration += 24
    return round(duration, 2)


def _ensure_unique_sheet_name(base: str, used: set[str]) -> str:
    name = _sanitize_sheet_name(base)
    if name not in used:
        used.add(name)
        return name
    suffix = 2
    while True:
        candidate = _sanitize_sheet_name(f"{name}_{suffix}")
        if candidate not in used:
            used.add(candidate)
            return candidate
        suffix += 1


def _add_group_sheets(
    wb: Workbook,
    per_group_rows: Dict[str, List[List[str]]],
    per_group_hours: Dict[str, float],
) -> None:
    used_names = {ws.title for ws in wb.worksheets}
    headers = [
        "Дата",
        "День недели",
        "№ пары",
        "Время",
        "Дисциплина",
        "Аудитория",
        "Комментарий",
    ]
    for group_name in sorted(per_group_rows.keys()):
        rows = _sort_group_rows(per_group_rows[group_name])
        if not rows:
            continue
        sheet_title = _ensure_unique_sheet_name(group_name, used_names)
        ws = wb.create_sheet(title=sheet_title)
        ws.append([f"Группа: {group_name}", "", "", "", "", "", ""])
        ws.append(headers)
        header_row = ws[2]
        header_fill = PatternFill("solid", fgColor="E3F2FD")
        thin_side = Side(style="thin", color="CCCCCC")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A3"
        _auto_fit_columns(ws, max_width=35)
        total_hours = per_group_hours.get(group_name, 0.0)
        if total_hours:
            ws.append([])
            ws.append(["", "", "", "", "", "", f"Итого: {total_hours:.1f} часов"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)


def _add_teacher_sheets(
    wb: Workbook,
    per_teacher_rows: Dict[str, List[List[str]]],
    per_teacher_hours: Dict[str, float],
) -> None:
    used_names = {ws.title for ws in wb.worksheets}
    headers = [
        "Дата",
        "День недели",
        "№ пары",
        "Время",
        "Дисциплина",
        "Преподаватель",
        "Аудитория",
        "Комментарий",
    ]
    for teacher_name in sorted(per_teacher_rows.keys()):
        rows = per_teacher_rows[teacher_name]
        if not rows:
            continue
        sheet_title = _ensure_unique_sheet_name(teacher_name, used_names)
        ws = wb.create_sheet(title=sheet_title)
        ws.append([f"Преподаватель: {teacher_name}", "", "", "", "", "", "", ""])
        ws.append(headers)
        header_row = ws[2]
        header_fill = PatternFill("solid", fgColor="FFF3E0")
        thin_side = Side(style="thin", color="CCCCCC")
        border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        for cell in header_row:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border
        for row in rows:
            ws.append(row)
        ws.freeze_panes = "A3"
        _auto_fit_columns(ws, max_width=40)
        total_hours = per_teacher_hours.get(teacher_name, 0.0)
        if total_hours:
            ws.append([])
            ws.append(["", "", "", "", "", "", "", f"Итого: {total_hours:.1f} часов"])
            for cell in ws[ws.max_row]:
                cell.font = Font(bold=True)


def create_group_workbook(
    group_name: str,
    rows: List[List[str]],
    teacher_name: str,
    semester_label: str,
    total_hours: float,
) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Расписание"
    ws.merge_cells("A1:H1")
    ws["A1"] = f"Группа: {group_name}"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.merge_cells("A2:H2")
    ws["A2"] = f"Преподаватель: {teacher_name} | {semester_label}"
    ws["A2"].alignment = Alignment(horizontal="center")

    headers = [
        "Дата",
        "День недели",
        "№ пары",
        "Время",
        "Дисциплина",
        "Преподаватель",
        "Аудитория",
        "Комментарий",
    ]
    ws.append(headers)
    header_row = ws[3]
    header_fill = PatternFill("solid", fgColor="E0F2F1")
    thin_side = Side(style="thin", color="CCCCCC")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    for cell in header_row:
        cell.font = Font(bold=True)
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = border

    for row in rows:
        data_row = row[:5] + [teacher_name] + row[5:]
        ws.append(data_row)
    ws.freeze_panes = "A4"
    _auto_fit_columns(ws, max_width=35)
    if total_hours:
        ws.append([])
        ws.append(["", "", "", "", "", "", "", f"Итого: {total_hours:.1f} часов"])
        for cell in ws[ws.max_row]:
            cell.font = Font(bold=True)
    return wb


def build_group_archive_bytes(
    per_group_rows: Dict[str, List[List[str]]],
    per_group_hours: Dict[str, float],
    teacher_name: str,
    semester_label: str,
) -> Tuple[Optional[bytes], int]:
    buffer = io.BytesIO()
    added = 0
    with zipfile.ZipFile(buffer, "w", zipfile.ZIP_DEFLATED) as archive:
        for group_name in sorted(per_group_rows.keys()):
            rows = _sort_group_rows(per_group_rows[group_name])
            if not rows:
                continue
            total_hours = per_group_hours.get(group_name, 0.0)
            wb = create_group_workbook(group_name, rows, teacher_name, semester_label, total_hours)
            temp_stream = io.BytesIO()
            wb.save(temp_stream)
            archive.writestr(f"{_sanitize_filename(group_name)}.xlsx", temp_stream.getvalue())
            added += 1
    if added == 0:
        return None, 0
    buffer.seek(0)
    return buffer.getvalue(), added


def _save_group_workbooks(
    per_group_rows: Dict[str, List[List[str]]],
    teacher_name: str,
    semester_label: str,
    output_dir: Path,
    base_stem: str,
    per_group_hours: Dict[str, float],
) -> List[Path]:
    group_dir = output_dir / f"{_sanitize_filename(base_stem)}_groups"
    group_dir.mkdir(parents=True, exist_ok=True)
    saved_files: List[Path] = []
    for group_name in sorted(per_group_rows.keys()):
        rows = _sort_group_rows(per_group_rows[group_name])
        if not rows:
            continue
        total_hours = per_group_hours.get(group_name, 0.0)
        wb = create_group_workbook(group_name, rows, teacher_name, semester_label, total_hours)
        filename = f"{_sanitize_filename(group_name)}.xlsx"
        saved_files.append(save_workbook(wb, group_dir, filename))
    return saved_files
def _sort_group_rows(rows: List[List[str]]) -> List[List[str]]:
    def parse_date(value: str) -> dt.date:
        try:
            return dt.datetime.strptime(value, "%d.%m.%Y").date()
        except Exception:
            return dt.date.min

    def parse_time(value: str) -> dt.time:
        match = re.search(r"(\d{1,2})[:.](\d{2})", value or "")
        if match:
            hour, minute = map(int, match.groups())
            return dt.time(hour=hour, minute=minute)
        return dt.time.min

    return sorted(
        rows,
        key=lambda row: (
            row[4] or "",
            parse_date(row[0]),
            parse_time(row[3]),
        ),
    )



def save_workbook(wb: Workbook, output_dir: Path, filename: Optional[str]) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    base_path = Path(filename) if filename else Path(f"{dt.datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    stem = base_path.stem
    attempt = 0
    while True:
        suffix = f"_{attempt}" if attempt else ""
        target = output_dir / f"{stem}{suffix}.xlsx"
        try:
            wb.save(target)
            return target
        except PermissionError:
            attempt += 1
            if attempt > 5:
                raise


async def send_via_telegram(token: str, chat_id: str, file_path: Path, caption: str) -> None:
    if Bot is None:
        raise RuntimeError("Модуль python-telegram-bot недоступен. Установите зависимость.")
    bot = Bot(token=token)
    async with bot:
        try:
            with file_path.open("rb") as fh:
                await bot.send_document(chat_id=chat_id, document=fh, filename=file_path.name, caption=caption)
            logger.info("📤 Файл отправлен в Telegram (chat_id=%s)", chat_id)
        except TelegramError as exc:  # pragma: no cover
            logger.error("Не удалось отправить файл: %s", exc)


async def main_async() -> None:
    args = parse_args()
    if args.silent:
        logging.getLogger().setLevel(logging.WARNING)

    start_date, end_date, semester_label = resolve_semester_bounds(args.semester, args.year, args.start, args.end)
    entity_type = API_TYPE_GROUP if args.mode == "student" else API_TYPE_TEACHER

    timetable = await fetch_semester_schedule(args.query, entity_type, start_date, end_date)
    if not timetable:
        logger.warning("⚠️ За указанный период не найдено расписания.")
        return

    (
        workbook,
        per_group_rows,
        _,
        total_hours,
        per_group_hours,
        _,
    ) = build_excel_workbook(args.query, args.mode, semester_label, timetable)
    output_dir = Path(args.output)
    filename = args.filename or f"{_sanitize_sheet_name(args.query)}_{semester_label.replace(' ', '_')}.xlsx"
    file_path = save_workbook(workbook, output_dir, filename)
    logger.info("💾 Файл сохранён: %s (суммарно %.1f часов)", file_path, total_hours)

    if args.mode == "teacher" and args.split_groups and per_group_rows:
        base_stem = Path(file_path).stem
        split_files = _save_group_workbooks(
            per_group_rows,
            args.query,
            semester_label,
            output_dir,
            base_stem,
            per_group_hours,
        )
        logger.info("📚 Дополнительно сохранено %d файлов по группам в %s", len(split_files), output_dir / f"{_sanitize_filename(base_stem)}_groups")

    if args.send:
        token = args.bot_token or CONFIG_BOT_TOKEN
        chat_id = args.chat_id or CONFIG_CHAT_ID
        if not token or not chat_id:
            logger.error("Не удалось отправить файл: не указан бот-токен или chat_id.")
            return
        await send_via_telegram(str(token), str(chat_id), file_path, f"Экспорт расписания: {args.query}")


def main() -> None:
    asyncio.run(main_async())


if __name__ == "__main__":
    main()

